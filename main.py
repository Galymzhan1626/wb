"""
FFCalculator — калькулятор себестоимости поставок для Wildberries и Ozon.

Структура:
  - Конфигурация магазинов и настроек
  - Авторизация в приложении (streamlit-authenticator)
  - Загрузка прайсов из Google Sheets
  - Работа с Wildberries API (заказы по поставке)
  - Парсинг PDF-накладных Ozon
  - Генерация ЭСФ из уведомления о выкупе WB
  - UI (вкладки по магазину/маркетплейсу)
"""

import os
import time
import codecs
import base64
import json
from io import BytesIO

import requests
import pdfplumber
import openpyxl
import pandas as pd
import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
import streamlit_authenticator as stauth


# =========================================================
# КОНФИГУРАЦИЯ
# =========================================================

DEFAULT_FF_COST = 400

WB_SHOPS = [
    "Тлеубаева", "Bonitas", "Мамутова", "Тастанов", "Bastau", "Шукурова",
    "Диханбаев", "Diamond", "Хаким", "Fariza", "Aibar", "Байпакова",
    "Абеденов", "Махамбетова", "Кыдырова", "Жораев",
]
WB_SHOPS_WITHOUT_FF = ["Диханбаев", "Хаким", "Diamond"]

OZON_SHOPS = [
    "Магазин 1",  # Сюда можно добавить другие магазины Ozon
]

ALL_SHOPS = WB_SHOPS + OZON_SHOPS

GSHEETS_SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]

# Соответствие "магазин в интерфейсе" -> "ключ в st.secrets['wb_api_keys']"
# Если магазина нет в этом словаре, вкладка WB API будет недоступна для него.
WB_SHOP_TO_SECRET_KEY = {
    "Абеденов": "Абеденов",
}


st.set_page_config(page_title="Калькулятор Поставок", layout="centered", page_icon="📦")

st.markdown("""
    <style>
    .stTable {font-size: 14px;}
    .reportview-container .main .block-container {padding-top: 2rem;}
    </style>
    """, unsafe_allow_html=True)


# =========================================================
# АВТОРИЗАЦИЯ
# =========================================================

def get_auth_credentials():
    """
    Логин/пароль берём из st.secrets, если они там заданы (рекомендуется),
    иначе используем встроенные значения как запасной вариант.
    Хранить пароль прямо в коде небезопасно — лучше вынести в secrets:

        [auth]
        username = "SeiE003YAN8J"
        name = "Менеджер"
        password_hash = "$2b$12$..."
    """
    auth_secrets = st.secrets.get("auth", {})
    username = auth_secrets.get("username", "SeiE003YAN8J")
    name = auth_secrets.get("name", "Менеджер")
    password_hash = auth_secrets.get(
        "password_hash",
        "$2b$12$OsSAaw38p2ICx2Xj3Yct6u.OnnwqaW99obBa1IcoTvi8GvIEbWnSa",
    )
    return {"usernames": {username: {"name": name, "password": password_hash}}}


authenticator = stauth.Authenticate(
    get_auth_credentials(),
    "delivery_app",
    st.secrets.get("auth", {}).get("cookie_key", "super_secret_key_xyz_123"),
    cookie_expiry_days=7,
)

authenticator.login()

if st.session_state.get("authentication_status") is False:
    st.error("❌ Неверный логин или пароль")
    st.stop()

if st.session_state.get("authentication_status") is None:
    st.warning("Введите логин и пароль")
    st.stop()

authenticator.logout("Выйти", "sidebar")
st.sidebar.write(f"👤 {st.session_state.get('name')}")

st.title("📦 Система расчёта себестоимости")
st.markdown("---")


# =========================================================
# GOOGLE SHEETS
# =========================================================

def get_google_credentials() -> Credentials:
    """
    Источники учётных данных, в порядке приоритета:
      1. st.secrets["gcp_service_account"]  (рекомендуется для облака)
      2. локальный файл creds.json рядом со скриптом (для локальной разработки)
    """
    if "gcp_service_account" in st.secrets:
        service_account_info = dict(st.secrets["gcp_service_account"])
        return Credentials.from_service_account_info(service_account_info, scopes=GSHEETS_SCOPES)

    local_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "creds.json")
    if os.path.exists(local_path):
        return Credentials.from_service_account_file(local_path, scopes=GSHEETS_SCOPES)

    raise RuntimeError(
        "Не найдены учётные данные Google. Добавь [gcp_service_account] в st.secrets "
        "или файл creds.json рядом со скриптом."
    )


@st.cache_data(ttl=300)
def load_prices_from_gsheets(shop_name: str):
    try:
        creds = get_google_credentials()
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_url(st.secrets["sheet_url"])
        worksheet = spreadsheet.worksheet(shop_name)
        df = pd.DataFrame(worksheet.get_all_records())
        return df, None
    except gspread.exceptions.WorksheetNotFound:
        return None, f"На листе Google Sheets нет вкладки «{shop_name}»."
    except Exception as e:
        return None, f"Ошибка доступа к Google Sheets: {e}"


# =========================================================
# WILDBERRIES API
# =========================================================

def decode_jwt_payload(token: str) -> dict | None:
    """Декодирует payload JWT без проверки подписи — только чтобы посмотреть exp/scope."""
    try:
        payload_b64 = token.split(".")[1]
        payload_b64 += "=" * (-len(payload_b64) % 4)
        return json.loads(base64.urlsafe_b64decode(payload_b64))
    except Exception:
        return None


def wb_token_diagnostics(api_key: str) -> str | None:
    """Возвращает предупреждение, если с токеном что-то похоже не так, иначе None."""
    payload = decode_jwt_payload(api_key)
    if not payload:
        return None
    exp = payload.get("exp")
    if exp:
        remaining_days = (exp - time.time()) / 86400
        if remaining_days < 0:
            return "⏰ Срок действия токена WB истёк. Сгенерируй новый ключ в ЛК WB."
        if remaining_days < 3:
            return f"⏰ Токен WB истекает через {remaining_days:.1f} дн. — стоит перевыпустить заранее."
    return None


@st.cache_data(ttl=60)
def get_supply_orders(supply_id: str, api_key: str):
    """
    Возвращает (summary_df, error_message).
    error_message пишется по-человечески, с учётом типичных причин 401/403/404.
    """
    clean_id = supply_id.strip()
    clean_key = api_key.strip()
    headers = {"Authorization": clean_key}
    url_direct = f"https://marketplace-api.wildberries.ru/api/v3/supplies/{clean_id}/orders"

    try:
        res = requests.get(url_direct, headers=headers, timeout=15)

        if res.status_code == 200:
            orders = res.json().get("orders", [])
            if orders:
                df = pd.DataFrame(orders)
                summary = df["article"].value_counts().reset_index()
                summary.columns = ["Артикул", "Заказ (уп)"]
                return summary, None
            # 200, но пустой список — пробуем общий эндпоинт ниже

        elif res.status_code == 401:
            return None, (
                "401 Unauthorized — WB не принял токен. Возможные причины:\n"
                "- у токена не включена категория доступа «Маркетплейс» (Поставки/Сборочные задания);\n"
                "- значение токена в secrets повреждено (лишние пробелы/переносы/обрезка);\n"
                "- токен отозван или перевыпущен в ЛК WB.\n"
                "Проверь ЛК WB → Настройки → Доступ к API."
            )
        elif res.status_code == 403:
            return None, "403 Forbidden — токен валиден, но нет прав на этот конкретный ресурс/магазин."

        url_all = "https://marketplace-api.wildberries.ru/api/v3/orders"
        params = {"limit": 1000, "next": 0}
        res_all = requests.get(url_all, headers=headers, params=params, timeout=15)

        if res_all.status_code == 200:
            all_orders = res_all.json().get("orders", [])
            filtered = [o for o in all_orders if str(o.get("supplyId")) == clean_id]
            if filtered:
                df = pd.DataFrame(filtered)
                summary = df["article"].value_counts().reset_index()
                summary.columns = ["Артикул", "Заказ (уп)"]
                return summary, None
            return None, f"Заказы для поставки {clean_id} не найдены."

        if res_all.status_code == 401:
            return None, "401 Unauthorized на общем эндпоинте заказов — проблема в самом токене (см. выше)."

        return None, f"Ошибка API: {res_all.status_code} — {res_all.text[:300]}"

    except requests.exceptions.Timeout:
        return None, "Таймаут запроса к WB API. Попробуй ещё раз чуть позже."
    except requests.exceptions.ConnectionError as e:
        return None, f"Не удалось подключиться к WB API: {e}"
    except Exception as e:
        return None, f"Ошибка: {str(e)}"


# =========================================================
# OZON PDF ПАРСЕР
# =========================================================

def parse_ozon_pdf(file) -> pd.DataFrame:
    items = []
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    if row and len(row) >= 5:
                        article = str(row[3]).strip() if row[3] else ""
                        qty_raw = str(row[4]).strip() if row[4] else ""
                        if article in ("", "Артикул", "None"):
                            continue
                        try:
                            qty = int(qty_raw)
                        except ValueError:
                            qty = 1
                        items.append({"Артикул": article, "Заказ (уп)": qty})

    if not items:
        return pd.DataFrame()

    df = pd.DataFrame(items)
    return df.groupby("Артикул")["Заказ (уп)"].sum().reset_index()


# =========================================================
# ОБЩИЙ РАСЧЁТ И ВЫВОД
# =========================================================

def show_results(summary: pd.DataFrame, df_prices: pd.DataFrame, selected_shop: str, current_ff_rate: float):
    res = pd.merge(
        summary,
        df_prices[["Артикул", "Количество в упаковке", "Цена за штуку"]],
        on="Артикул",
        how="left",
    )

    unmatched = res[res["Цена за штуку"].isna()]["Артикул"].tolist()
    if unmatched:
        st.warning(f"⚠️ **{len(unmatched)} SKU** не найдены в прайсе и пропущены:\n{', '.join(map(str, unmatched))}")

    res = res.dropna(subset=["Цена за штуку"])

    if res.empty:
        st.error("❌ Нет данных для расчета. Проверьте артикулы.")
        return

    res["Всего шт"] = res["Заказ (уп)"] * res["Количество в упаковке"]
    res["Цена товара"] = res["Всего шт"] * res["Цена за штуку"]

    st.subheader("📊 Результаты расчёта")
    st.dataframe(
        res[["Артикул", "Заказ (уп)", "Всего шт", "Цена за штуку", "Цена товара"]].style.format({
            "Цена товара": "{:,.0f} ₸",
            "Цена за штуку": "{:,.0f} ₸",
            "Всего шт": "{:,.0f}",
            "Заказ (уп)": "{:,.0f}",
        }),
        use_container_width=True,
        hide_index=True,
    )

    total_packs = res["Заказ (уп)"].sum()
    total_items_cost = res["Цена товара"].sum()
    total_ff = total_packs * current_ff_rate
    grand_total = total_items_cost + total_ff

    st.markdown("---")
    c_res1, c_res2 = st.columns(2)
    with c_res1:
        st.write(f"📦 **Заказов:** {total_packs} уп.")
        st.write(f"⚙️ **Фулфилмент:** {total_ff:,.0f} ₸")
        st.write(f"💰 **Стоимость товара:** {total_items_cost:,.0f} ₸")
    with c_res2:
        st.metric(label="ИТОГО К ОПЛАТЕ", value=f"{grand_total:,.0f} ₸")

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            res.to_excel(writer, index=False, sheet_name="Расчет")
        st.download_button(
            "⬇️ Скачать Excel",
            data=output.getvalue(),
            file_name=f"Расчет_{selected_shop}_{time.strftime('%d%m')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


def build_esf_txt(items: list[dict]) -> bytes:
    """Формирует .txt для портала ЭСФ в кодировке UTF-16 LE с BOM."""
    txt_lines = [
        "Раздел G. Данные\u00a0по\u00a0товарам, работам,\u00a0услугам" + "\t" * 17,
        "\t" * 17,
        (
            "Признак происхождения товара, работ, услуг*\t"
            "Наименование товаров, работ, услуг*\t"
            "Наименование товаров в соответствии с Декларацией на товары или заявлением о ввозе товаров и уплате косвенных налогов\t"
            "Код товара (ТН ВЭД ЕАЭС)\tЕд. изм.\tКол-во (объем)\t"
            "Цена (тариф) за единицу товара, работы, услуги без косвенных налогов\t"
            "Стоимость товаров, работ, услуг без косвенных налогов*\t"
            "Акциз- Ставка\tАкциз- Сумма\t"
            "Размер оборота по реализации (облагаемый/необлагаемый оборот)*\t"
            "НДС- Ставка\tНДС- Сумма\t"
            "Стоимость товаров, работ, услуг с учетом косвенных налогов*\t"
            "№ Декларации на товары, заявления в рамках ТС, СТ-1 или СТ-KZ\t"
            "Номер товарной позиции из заявления в рамках ТС или Декларации на товары\t"
            "Идентификатор товара, работ, услуг\tДополнительные данные"
        ),
        "\t" * 17,
        "2\t3\t3/1\t4\t5\t6\t7\t8\t9\t10\t11\t12\t13\t14\t15\t16\t17\t18",
    ]

    for item in items:
        qty_val = item["qty"]
        qty_str = (
            str(int(qty_val))
            if isinstance(qty_val, float) and qty_val.is_integer()
            else f"{qty_val:.2f}".replace(".", ",")
        )
        price_str = f"{item['price']:.2f}".replace(".", ",")
        sum_str = f"{item['sum']:.2f}".replace(".", ",")

        row_fields = [
            "5", str(item["name"]), "", item["tnved"], "Штука",
            qty_str, price_str, sum_str, "", "",
            sum_str, "Без НДС", "0", sum_str, "", "", "1", "",
        ]
        txt_lines.append("\t".join(row_fields))

    txt_content = "\r\n".join(txt_lines)
    return codecs.BOM_UTF16_LE + txt_content.encode("utf-16-le")


def parse_wb_notice(file_wb, df_prices: pd.DataFrame) -> list[dict]:
    tnved_dict = {}
    if "Артикул" in df_prices.columns and "ТН ВЭД" in df_prices.columns:
        df_prices["Артикул_clean"] = df_prices["Артикул"].astype(str).str.strip()
        tnved_dict = pd.Series(df_prices["ТН ВЭД"].values, index=df_prices["Артикул_clean"]).to_dict()

    wb_source = openpyxl.load_workbook(file_wb, data_only=True)
    ws_source = wb_source.active

    title = ws_source.cell(row=3, column=1).value
    if title:
        st.caption(f"Документ: **{title}**")

    items = []
    for row in ws_source.iter_rows(min_row=11, values_only=True):
        if not isinstance(row[0], int):
            continue

        raw_art = str(row[1]).strip() if row[1] is not None else ""
        raw_qty = row[3]
        raw_sum = row[4]

        try:
            qty = float(str(raw_qty).replace(" ", "").replace(",", ".")) if raw_qty is not None else 0
            total_sum = float(str(raw_sum).replace(" ", "").replace(",", ".")) if raw_sum is not None else 0
        except Exception:
            qty = raw_qty or 0
            total_sum = raw_sum or 0

        price_per_item = total_sum / qty if qty > 0 else 0

        item_tnved = tnved_dict.get(raw_art, "")
        if pd.isna(item_tnved):
            item_tnved = ""
        tnved_str = str(item_tnved).strip()
        if tnved_str.endswith(".0"):
            tnved_str = tnved_str[:-2]
        if tnved_str == "nan":
            tnved_str = ""

        items.append({
            "article": raw_art,
            "name": row[2],
            "qty": qty,
            "sum": total_sum,
            "price": price_per_item,
            "tnved": tnved_str,
        })

    return items


# =========================================================
# UI
# =========================================================

col_main, col_refresh = st.columns([4, 1])
with col_main:
    selected_shop = st.selectbox("🎯 Выберите магазин:", ALL_SHOPS)
with col_refresh:
    st.markdown("<div style='margin-top: 28px'>", unsafe_allow_html=True)
    if st.button("🔄", help="Обновить прайс из Google Sheets"):
        st.cache_data.clear()
        st.rerun()

is_ozon_shop = selected_shop in OZON_SHOPS
if is_ozon_shop:
    current_ff_rate = DEFAULT_FF_COST
else:
    current_ff_rate = 0 if selected_shop in WB_SHOPS_WITHOUT_FF else DEFAULT_FF_COST

with st.spinner("⏳ Синхронизация с Google Sheets..."):
    df_prices, error = load_prices_from_gsheets(selected_shop)

if error:
    st.error(f"❌ {error}")
    st.stop()

if df_prices is None or df_prices.empty:
    st.error("❌ Прайс не загружен — получен пустой результат из Google Sheets")
    st.stop()

st.caption(f"✅ Прайс обновлен в {time.strftime('%H:%M')} | {len(df_prices)} SKU")
st.subheader(f"🚚 Поставка: {selected_shop} ({'Ozon' if is_ozon_shop else 'Wildberries'})")

if is_ozon_shop:
    tab_ozon = st.tabs(["🔵 Загрузка Ozon PDF"])[0]
    with tab_ozon:
        ozon_file = st.file_uploader("Загрузите PDF от Ozon", type=["pdf"], key="ozon_pdf")
        if ozon_file:
            with st.spinner("Читаем PDF..."):
                summary_ozon = parse_ozon_pdf(ozon_file)
            if summary_ozon.empty:
                st.error("❌ Не удалось извлечь артикулы из PDF. Проверьте формат файла.")
            else:
                st.success(f"✅ Найдено позиций: {len(summary_ozon)}")
                show_results(summary_ozon, df_prices, selected_shop, current_ff_rate)

else:
    tab_api, tab_file, tab_esf = st.tabs(["🔗 Wildberries API", "📂 Загрузка Excel", "📝 ЭСФ"])

    # ---------------- ВКЛАДКА 1: WILDBERRIES API ----------------
    with tab_api:
        secret_key_name = WB_SHOP_TO_SECRET_KEY.get(selected_shop)
        api_key = None
        if secret_key_name:
            raw_key = st.secrets.get("wb_api_keys", {}).get(secret_key_name)
            if raw_key:
                api_key = raw_key.strip()

        if not api_key:
            st.info(
                "ℹ️ Для этого магазина API-ключ WB не настроен в secrets "
                "(`wb_api_keys` → соответствующий ключ). Используй вкладку «Загрузка Excel»."
            )
        else:
            warning = wb_token_diagnostics(api_key)
            if warning:
                st.warning(warning)

            supply_id = st.text_input(
                "Номер поставки WB", key="wb_supply_id", placeholder="Например: WB-GI-123456789"
            )
            if st.button("📥 Получить заказы по поставке", key="fetch_supply_btn"):
                if not supply_id.strip():
                    st.warning("⚠️ Введите номер поставки.")
                else:
                    with st.spinner("Запрашиваем данные у Wildberries..."):
                        summary_api, api_error = get_supply_orders(supply_id, api_key)
                    if api_error:
                        st.error(f"❌ {api_error}")
                    else:
                        st.success(f"✅ Найдено позиций: {len(summary_api)}")
                        show_results(summary_api, df_prices, selected_shop, current_ff_rate)

    # ---------------- ВКЛАДКА 2: ЗАГРУЗКА EXCEL ----------------
    with tab_file:
        delivery_file = st.file_uploader("Файл поставки (колонка F)", type=["xlsx"], key="delivery_upload")
        if delivery_file:
            try:
                df_raw = pd.read_excel(delivery_file, skiprows=4, usecols="F").dropna()
                df_raw.columns = ["Артикул"]
                summary_file = df_raw["Артикул"].value_counts().reset_index()
                summary_file.columns = ["Артикул", "Заказ (уп)"]
                show_results(summary_file, df_prices, selected_shop, current_ff_rate)
            except Exception as e:
                st.error(f"❌ Ошибка файла: {e}")

    # ---------------- ВКЛАДКА 3: ЭСФ ----------------
    with tab_esf:
        st.subheader("Генерация ЭСФ из уведомления о выкупе WB")

        file_wb = st.file_uploader(
            "Загрузите Excel (уведомление о выкупе WB)", type=["xlsx"], key="wb_notice"
        )

        if file_wb:
            try:
                items = parse_wb_notice(file_wb, df_prices)
            except Exception as e:
                st.error(f"🔴 Ошибка при обработке уведомления: {e}")
                items = []

            if file_wb and not items:
                st.warning("⚠️ Товары не найдены. Проверьте формат уведомления.")
            elif items:
                st.success(f"✅ Найдено позиций: {len(items)}")
                st.markdown("---")
                st.write("**Выберите формат для скачивания:**")

                encoded_content = build_esf_txt(items)
                st.download_button(
                    label="📄 Скачать ЭСФ (.txt) — Рекомендуется",
                    data=encoded_content,
                    file_name=f"ЭСФ_{selected_shop}_{time.strftime('%d%m')}.txt",
                    mime="text/plain",
                    key="download_txt_btn",
                )

                st.caption("или")

                BASE_DIR = os.path.dirname(os.path.abspath(__file__))
                TEMPLATE_ESF_FILE = os.path.join(BASE_DIR, "schedule.xlsx")

                if not os.path.exists(TEMPLATE_ESF_FILE):
                    st.info("💡 Скачивание в Excel недоступно — файл 'schedule.xlsx' не найден в папке проекта.")
                else:
                    try:
                        wb_template = openpyxl.load_workbook(TEMPLATE_ESF_FILE)
                        ws_template = wb_template.active

                        START_ROW = 6
                        for i, item in enumerate(items):
                            row_idx = START_ROW + i
                            ws_template.cell(row=row_idx, column=1, value=5)
                            ws_template.cell(row=row_idx, column=2, value=item["name"])
                            ws_template.cell(row=row_idx, column=3, value="")
                            ws_template.cell(row=row_idx, column=4, value=item["tnved"])
                            ws_template.cell(row=row_idx, column=5, value="Штука")
                            ws_template.cell(row=row_idx, column=6, value=item["qty"])
                            ws_template.cell(row=row_idx, column=7, value=item["price"])
                            ws_template.cell(row=row_idx, column=8, value=item["sum"])
                            ws_template.cell(row=row_idx, column=9, value="")
                            ws_template.cell(row=row_idx, column=10, value="")
                            ws_template.cell(row=row_idx, column=11, value=item["sum"])
                            ws_template.cell(row=row_idx, column=12, value="Без НДС")
                            ws_template.cell(row=row_idx, column=13, value=0)
                            ws_template.cell(row=row_idx, column=14, value=item["sum"])
                            ws_template.cell(row=row_idx, column=15, value="")
                            ws_template.cell(row=row_idx, column=16, value="")
                            ws_template.cell(row=row_idx, column=17, value=1)
                            ws_template.cell(row=row_idx, column=18, value="")

                        output_esf = BytesIO()
                        wb_template.save(output_esf)
                        output_esf.seek(0)

                        st.download_button(
                            label="⬇️ Скачать заполненный шаблон (.xlsx)",
                            data=output_esf.getvalue(),
                            file_name=f"ЭСФ_{selected_shop}_{time.strftime('%d%m')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_xlsx_btn",
                        )
                    except Exception as e:
                        st.error(f"Ошибка при сохранении в Excel-шаблон: {e}")
